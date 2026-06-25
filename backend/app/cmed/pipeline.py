"""
Pipeline de integração CMED/PMC — PharmaPrice
============================================
Responsabilidades:
  1. Fazer download do arquivo XLS/XLSX da ANVISA
  2. Parsear as colunas relevantes
  3. Inserir/atualizar os registros no banco PostgreSQL

Estrutura da planilha (confirmada em 10/06/2026):
  - Cabeçalho real: linha 43
  - Dados: linhas 44 em diante
  - Total: ~25.435 linhas x 74 colunas
  - PMC 18% (MG padrão): coluna 47 (índice base 0)
"""
import logging
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Optional

import httpx
import openpyxl
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

# URL pública da tabela de preços da ANVISA/CMED
CMED_URL = (
    "https://www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed/precos"
)

# Linha do cabeçalho real na planilha (1-indexed, conforme openpyxl)
HEADER_ROW = 43

# Mapeamento: nome da coluna no Excel → atributo no modelo
# Índices confirmados por inspeção direta do arquivo real
COLUMN_MAP = {
    0:  "substancia",
    1:  "cnpj",
    2:  "laboratorio",
    3:  "codigo_ggrem",
    4:  "registro",
    5:  "ean1",
    8:  "produto",
    9:  "apresentacao",
    10: "classe_terapeutica",
    11: "tipo_produto",
    12: "regime_preco",
    13: "pf_sem_impostos",
    14: "pf_0",
    39: "pmc_sem_impostos",
    40: "pmc_0",
    41: "pmc_12",
    43: "pmc_17",
    47: "pmc_18",   # MG padrão
    49: "pmc_19",
    53: "pmc_20",
    65: "restricao_hospitalar",
    66: "cap",
    67: "confaz_87",
    68: "icms_0",
    71: "comercializacao_2025",
    72: "tarja",
}


def _parse_decimal(value) -> Optional[Decimal]:
    """Converte string ou número para Decimal. Retorna None se inválido."""
    if value is None:
        return None
    try:
        # A planilha usa vírgula como separador decimal em alguns campos
        cleaned = str(value).strip().replace(",", ".").replace(" ", "")
        if cleaned in ("", "-", "    -     "):
            return None
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _parse_bool_flag(value) -> bool:
    """Interpreta flags como 'Sim', 'S', '1', True → True."""
    if value is None:
        return False
    return str(value).strip().upper() in ("SIM", "S", "1", "TRUE", "X")


def _extrair_data_publicacao(ws) -> str:
    """Extrai a data de publicação do cabeçalho da planilha (linha 3)."""
    try:
        valor = ws.cell(row=3, column=1).value or ""
        # Ex: "Publicada em 10/06/2026 13h30min."
        match = re.search(r"\d{2}/\d{2}/\d{4}.*", str(valor))
        return match.group(0).rstrip(".") if match else str(valor)
    except Exception:
        return ""


def _row_to_dict(row_values: tuple) -> dict:
    """Converte uma linha da planilha em dicionário para o modelo."""
    data = {}
    for col_idx, field_name in COLUMN_MAP.items():
        if col_idx >= len(row_values):
            continue
        value = row_values[col_idx]

        if field_name in ("pf_sem_impostos", "pf_0", "pmc_sem_impostos",
                          "pmc_0", "pmc_12", "pmc_17", "pmc_18",
                          "pmc_19", "pmc_20"):
            data[field_name] = _parse_decimal(value)
        elif field_name in ("restricao_hospitalar", "cap", "confaz_87", "icms_0"):
            data[field_name] = _parse_bool_flag(value)
        else:
            # Campo texto — limpar espaços e valores "-"
            cleaned = str(value).strip() if value is not None else None
            if cleaned in ("", "-", "    -     ", "None"):
                cleaned = None
            data[field_name] = cleaned

    return data


def parsear_xlsx(conteudo: bytes) -> tuple[list[dict], str]:
    """
    Lê o conteúdo binário do XLSX da CMED e retorna:
      - lista de dicionários (um por medicamento)
      - data de publicação extraída do cabeçalho
    """
    wb = openpyxl.load_workbook(
        BytesIO(conteudo),
        read_only=True,
        data_only=True
    )
    ws = wb.active

    data_publicacao = _extrair_data_publicacao(ws)
    medicamentos = []
    encontrou_header = False
    header_row_idx = 0

    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Pular até a linha do cabeçalho real
        if row_num < HEADER_ROW:
            continue
        if row_num == HEADER_ROW:
            encontrou_header = True
            header_row_idx = row_num
            continue  # pula o cabeçalho

        # Ignorar linhas completamente vazias
        if all(v is None for v in row):
            continue

        # Ignorar linha se não tiver pelo menos SUBSTÂNCIA e PRODUTO
        if row[0] is None and row[8] is None:
            continue

        record = _row_to_dict(row)

        # Validação mínima: precisa de codigo_ggrem (chave única)
        if not record.get("codigo_ggrem"):
            continue

        medicamentos.append(record)

    wb.close()
    logger.info(
        f"Parsing concluído: {len(medicamentos)} medicamentos extraídos "
        f"(publicação: {data_publicacao})"
    )
    return medicamentos, data_publicacao


def parsear_arquivo_local(caminho: str) -> tuple[list[dict], str]:
    """
    Variante para desenvolvimento: lê de um arquivo local em vez de baixar.
    Útil para testes sem depender da URL da ANVISA.
    """
    with open(caminho, "rb") as f:
        conteudo = f.read()
    return parsear_xlsx(conteudo)


def salvar_no_banco(
    session: Session,
    medicamentos: list[dict],
    data_publicacao: str,
    fonte_url: str,
) -> dict:
    """
    Insere ou atualiza registros no banco.
    Usa UPSERT por codigo_ggrem (chave única).
    Retorna estatísticas da operação.
    """
    # Import aqui para evitar dependência circular
    from app.db.models import Medicamento

    inseridos = 0
    atualizados = 0
    erros = 0
    data_coleta = datetime.utcnow()

    for record in medicamentos:
        try:
            codigo_ggrem = record.get("codigo_ggrem")
            existente = session.query(Medicamento).filter_by(
                codigo_ggrem=codigo_ggrem
            ).first()

            if existente:
                # Atualizar campos
                for campo, valor in record.items():
                    setattr(existente, campo, valor)
                existente.data_coleta = data_coleta
                existente.fonte_url = fonte_url
                existente.data_publicacao_cmed = data_publicacao
                atualizados += 1
            else:
                novo = Medicamento(
                    **record,
                    data_coleta=data_coleta,
                    fonte_url=fonte_url,
                    data_publicacao_cmed=data_publicacao,
                )
                session.add(novo)
                inseridos += 1

            # Commit em lotes de 500 para não sobrecarregar a memória
            if (inseridos + atualizados) % 500 == 0:
                session.commit()
                logger.debug(
                    f"Progresso: {inseridos} inseridos, {atualizados} atualizados"
                )

        except Exception as e:
            logger.error(f"Erro ao processar GGREM {record.get('codigo_ggrem')}: {e}")
            session.rollback()
            erros += 1

    session.commit()

    stats = {
        "inseridos": inseridos,
        "atualizados": atualizados,
        "erros": erros,
        "total": inseridos + atualizados,
        "data_publicacao": data_publicacao,
    }
    logger.info(f"Banco atualizado: {stats}")
    return stats
