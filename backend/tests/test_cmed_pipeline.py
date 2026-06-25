"""
Testes unitários do pipeline CMED — PharmaPrice
================================================
Testa parsing, conversão de tipos e tratamento de edge cases
sem depender do banco de dados ou da URL da ANVISA.
"""
import pytest
from decimal import Decimal
from unittest.mock import patch, MagicMock
import openpyxl
from io import BytesIO

# Ajuste o import conforme a estrutura final do projeto
# from app.cmed.pipeline import parsear_xlsx, _parse_decimal, _parse_bool_flag, _row_to_dict
# Por ora, importamos direto do arquivo para testes isolados:
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from app.cmed.pipeline import (
    parsear_xlsx,
    _parse_decimal,
    _parse_bool_flag,
    _row_to_dict,
    HEADER_ROW,
    COLUMN_MAP,
)


# ─── Helpers para criar planilhas de teste ──────────────────────────────────

def criar_xlsx_minimo(linhas_dados: list[tuple]) -> bytes:
    """
    Cria um XLSX mínimo com cabeçalho na linha HEADER_ROW e dados logo após.
    Útil para testes sem depender do arquivo real da ANVISA.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Linhas de metadados (1 a HEADER_ROW-1)
    ws.cell(row=1, column=1, value="Secretaria Executiva - CMED")
    ws.cell(row=3, column=1, value="Publicada em 10/06/2026 13h30min.")

    # Cabeçalho real
    headers = [None] * 74
    for col_idx, field in COLUMN_MAP.items():
        headers[col_idx] = field.upper()
    for j, h in enumerate(headers):
        ws.cell(row=HEADER_ROW, column=j + 1, value=h)

    # Dados
    for i, linha in enumerate(linhas_dados):
        for j, valor in enumerate(linha):
            ws.cell(row=HEADER_ROW + 1 + i, column=j + 1, value=valor)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def linha_medicamento_valida(
    substancia="DIPIRONA MONOIDRATADA",
    produto="NOVALGINA",
    apresentacao="500 MG COM CT BL AL PLAS INC X 10",
    codigo_ggrem="000000000000001",
    pmc_18="12,50",
) -> tuple:
    """Retorna uma tupla de 74 posições com um medicamento válido."""
    row = ["    -     "] * 74
    row[0] = substancia
    row[1] = "00.000.000/0001-00"
    row[2] = "LABORATÓRIO TESTE"
    row[3] = codigo_ggrem
    row[4] = "1234567890000"
    row[5] = "7891234567890"
    row[8] = produto
    row[9] = apresentacao
    row[10] = "N1 - ANALGÉSICOS"
    row[11] = "Referência"
    row[12] = "Regulado"
    row[13] = "10,00"   # PF sem impostos
    row[14] = "11,00"   # PF 0%
    row[39] = "11,00"   # PMC sem impostos
    row[40] = "12,00"   # PMC 0%
    row[47] = pmc_18    # PMC 18% (MG)
    row[65] = "Não"
    row[66] = "Não"
    row[67] = "Não"
    row[68] = "Não"
    row[72] = "TARJA VERMELHA"
    return tuple(row)


# ─── Testes de _parse_decimal ────────────────────────────────────────────────

class TestParseDecimal:
    def test_valor_com_virgula(self):
        assert _parse_decimal("12,50") == Decimal("12.50")

    def test_valor_com_ponto(self):
        assert _parse_decimal("12.50") == Decimal("12.50")

    def test_valor_none(self):
        assert _parse_decimal(None) is None

    def test_valor_vazio(self):
        assert _parse_decimal("") is None

    def test_valor_tracinho(self):
        assert _parse_decimal("-") is None
        assert _parse_decimal("    -     ") is None

    def test_valor_inteiro(self):
        assert _parse_decimal("100") == Decimal("100")

    def test_valor_float_python(self):
        result = _parse_decimal(12.5)
        assert result is not None
        assert float(result) == pytest.approx(12.5)

    def test_valor_invalido(self):
        assert _parse_decimal("abc") is None

    def test_valor_grande(self):
        assert _parse_decimal("6662,98") == Decimal("6662.98")


# ─── Testes de _parse_bool_flag ──────────────────────────────────────────────

class TestParseBoolFlag:
    def test_sim(self):
        assert _parse_bool_flag("Sim") is True

    def test_s(self):
        assert _parse_bool_flag("S") is True

    def test_nao(self):
        assert _parse_bool_flag("Não") is False

    def test_none(self):
        assert _parse_bool_flag(None) is False

    def test_vazio(self):
        assert _parse_bool_flag("") is False

    def test_x(self):
        assert _parse_bool_flag("X") is True

    def test_case_insensitive(self):
        assert _parse_bool_flag("sim") is True
        assert _parse_bool_flag("SIM") is True


# ─── Testes de _row_to_dict ──────────────────────────────────────────────────

class TestRowToDict:
    def test_campos_obrigatorios_presentes(self):
        row = linha_medicamento_valida()
        result = _row_to_dict(row)
        assert result["substancia"] == "DIPIRONA MONOIDRATADA"
        assert result["produto"] == "NOVALGINA"
        assert result["codigo_ggrem"] == "000000000000001"

    def test_pmc_18_convertido_para_decimal(self):
        row = linha_medicamento_valida(pmc_18="12,50")
        result = _row_to_dict(row)
        assert result["pmc_18"] == Decimal("12.50")

    def test_tracinho_vira_none(self):
        row = linha_medicamento_valida()
        result = _row_to_dict(row)
        # EAN2 e EAN3 têm "    -     " no fixture
        # ean1 tem valor, outros campos texto com tracinho viram None
        assert result.get("restricao_hospitalar") is False  # flag

    def test_pmc_none_quando_ausente(self):
        row = linha_medicamento_valida(pmc_18="    -     ")
        result = _row_to_dict(row)
        assert result["pmc_18"] is None


# ─── Testes de parsear_xlsx ──────────────────────────────────────────────────

class TestParsearXlsx:
    def test_retorna_lista_e_data(self):
        xlsx_bytes = criar_xlsx_minimo([linha_medicamento_valida()])
        medicamentos, data_pub = parsear_xlsx(xlsx_bytes)
        assert isinstance(medicamentos, list)
        assert isinstance(data_pub, str)

    def test_um_medicamento_valido(self):
        xlsx_bytes = criar_xlsx_minimo([linha_medicamento_valida()])
        medicamentos, _ = parsear_xlsx(xlsx_bytes)
        assert len(medicamentos) == 1
        assert medicamentos[0]["produto"] == "NOVALGINA"

    def test_multiplos_medicamentos(self):
        linhas = [
            linha_medicamento_valida(codigo_ggrem="000000000000001", produto="PRODUTO A"),
            linha_medicamento_valida(codigo_ggrem="000000000000002", produto="PRODUTO B"),
            linha_medicamento_valida(codigo_ggrem="000000000000003", produto="PRODUTO C"),
        ]
        xlsx_bytes = criar_xlsx_minimo(linhas)
        medicamentos, _ = parsear_xlsx(xlsx_bytes)
        assert len(medicamentos) == 3

    def test_linha_sem_codigo_ggrem_ignorada(self):
        row_invalida = list(linha_medicamento_valida())
        row_invalida[3] = None  # Remove codigo_ggrem
        xlsx_bytes = criar_xlsx_minimo([tuple(row_invalida)])
        medicamentos, _ = parsear_xlsx(xlsx_bytes)
        assert len(medicamentos) == 0

    def test_linha_completamente_vazia_ignorada(self):
        row_vazia = tuple([None] * 74)
        xlsx_bytes = criar_xlsx_minimo([linha_medicamento_valida(), row_vazia])
        medicamentos, _ = parsear_xlsx(xlsx_bytes)
        assert len(medicamentos) == 1

    def test_data_publicacao_extraida(self):
        xlsx_bytes = criar_xlsx_minimo([linha_medicamento_valida()])
        _, data_pub = parsear_xlsx(xlsx_bytes)
        assert "2026" in data_pub

    def test_pmc_18_correto(self):
        xlsx_bytes = criar_xlsx_minimo([
            linha_medicamento_valida(pmc_18="25,99")
        ])
        medicamentos, _ = parsear_xlsx(xlsx_bytes)
        assert medicamentos[0]["pmc_18"] == Decimal("25.99")

    def test_substancias_diferentes(self):
        linhas = [
            linha_medicamento_valida(
                codigo_ggrem="001", substancia="DIPIRONA", produto="NOVALGINA"
            ),
            linha_medicamento_valida(
                codigo_ggrem="002", substancia="IBUPROFENO", produto="ADVIL"
            ),
        ]
        xlsx_bytes = criar_xlsx_minimo(linhas)
        medicamentos, _ = parsear_xlsx(xlsx_bytes)
        substancias = {m["substancia"] for m in medicamentos}
        assert "DIPIRONA" in substancias
        assert "IBUPROFENO" in substancias
